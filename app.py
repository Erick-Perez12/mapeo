"""
Mapeo por Ventana Geomecánica — Streamlit
==========================================
Requisitos:
    pip install streamlit pandas openpyxl

Ejecutar:
    streamlit run app.py
"""

import io
import json
from datetime import date

import pandas as pd
import streamlit as st

# ─────────────────────────────────────────────────────────────────────────────
# CATÁLOGOS
# ─────────────────────────────────────────────────────────────────────────────

TIPO_ESTRUCTURA = ["JN — Junta", "BED — Estrato", "F — Falla", "SZ — Cizalla", "CON — Contacto"]
TIPO_ESTRUCTURA_CODES = ["JN", "BED", "F", "SZ", "CON"]

FORMA_ESTRUCTURA = ["P — Plana", "C — Curva", "O — Ondulada", "E — Escalonada", "I — Irregular"]
FORMA_ESTRUCTURA_CODES = ["P", "C", "O", "E", "I"]

TERMINACION_OPTS = ["R — Roca", "D — Discontinuidad", "S — Sin observar"]
TERMINACION_CODES = ["R", "D", "S"]

ABERTURA_CAT = [
    {"label": "Masiva",                 "code": "0",       "mm": "0",      "r89": 6, "r76": 5},
    {"label": "Entre Abierta",          "code": "0.000001","mm": "<0.1",   "r89": 5, "r76": 4},
    {"label": "Abierta",                "code": "0.1",     "mm": "0.1–1",  "r89": 3, "r76": 3},
    {"label": "Muy Abierta",            "code": "1",       "mm": "1–5",    "r89": 1, "r76": 1},
    {"label": "Extremadamente Abierta", "code": "5",       "mm": ">5",     "r89": 0, "r76": 0},
]

CONTINUIDAD_CAT = [
    {"label": "<1 m",    "code": "0",  "r89": 6, "r76": 5},
    {"label": "1–3 m",   "code": "1",  "r89": 4, "r76": 4},
    {"label": "3–10 m",  "code": "3",  "r89": 2, "r76": 3},
    {"label": "10–20 m", "code": "10", "r89": 1, "r76": 1},
    {"label": ">20 m",   "code": "20", "r89": 0, "r76": 0},
]

ESPESOR_CAT = [
    {"label": "<5 mm", "code": "1"},
    {"label": ">5 mm", "code": "2"},
]

RELLENO_CAT = [
    {"code": "cwf", "tipo": 3, "sin89": 6, "sin76": 5},
    {"code": "si",  "tipo": 1, "dlt5_89": 4, "dlt5_76": 4, "dgt5_89": 2, "dgt5_76": 3},
    {"code": "sf",  "tipo": 1, "dlt5_89": 4, "dlt5_76": 4, "dgt5_89": 2, "dgt5_76": 3},
    {"code": "ep",  "tipo": 1, "dlt5_89": 4, "dlt5_76": 4, "dgt5_89": 2, "dgt5_76": 3},
    {"code": "ox",  "tipo": 1, "dlt5_89": 4, "dlt5_76": 4, "dgt5_89": 2, "dgt5_76": 3},
    {"code": "g",   "tipo": 2, "blt5_89": 2, "blt5_76": 2, "bgt5_89": 0, "bgt5_76": 0},
    {"code": "cl",  "tipo": 2, "blt5_89": 2, "blt5_76": 2, "bgt5_89": 0, "bgt5_76": 0},
    {"code": "ca",  "tipo": 2, "blt5_89": 2, "blt5_76": 2, "bgt5_89": 0, "bgt5_76": 0},
]
RELLENO_CODES = [""] + [r["code"] for r in RELLENO_CAT]

RUGOSIDAD_CAT = {
    "1": (6, 5), "2": (5, 4), "3": (5, 4),
    "4": (3, 2), "5": (3, 2),
    "6": (1, 0), "7": (1, 0),
    "8": (0, 0), "9": (0, 0),
}

ALTERACION_CAT = [
    {"code": "f",  "label": "f — Fresca",                    "r89": 6, "r76": 5},
    {"code": "d",  "label": "d — Débilmente meteorizada",    "r89": 5, "r76": 5},
    {"code": "m",  "label": "m — Moderadamente meteorizada", "r89": 3, "r76": 4},
    {"code": "a",  "label": "a — Altamente meteorizada",     "r89": 3, "r76": 3},
    {"code": "c",  "label": "c — Completamente meteorizada", "r89": 2, "r76": 2},
    {"code": "s",  "label": "s — Suelo residual",            "r89": 1, "r76": 1},
    {"code": "-1", "label": "-1 — Sin información",          "r89": 0, "r76": 0},
]
ALT_CODES = [""] + [a["code"] for a in ALTERACION_CAT]
ALT_LABELS = ["—"] + [a["label"] for a in ALTERACION_CAT]

JRC_RANGES = [
    "0–2 (perfil 9)", "2–4 (perfil 8)", "4–6 (perfil 7)", "6–8 (perfil 6)",
    "8–10 (perfil 5)", "10–12 (perfil 4)", "12–14 (perfil 3)",
    "14–16 (perfil 2)", "16–18 (perfil 1)", "18–20 (perfil 1)",
]

# ─────────────────────────────────────────────────────────────────────────────
# CÁLCULOS
# ─────────────────────────────────────────────────────────────────────────────

def get_relleno_rating(code: str, espesor_code: str):
    r = next((x for x in RELLENO_CAT if x["code"] == code), None)
    if not r:
        return None, None
    if r["tipo"] == 3:
        return r.get("sin89"), r.get("sin76")
    if r["tipo"] == 1:
        if espesor_code == "1":
            return r.get("dlt5_89"), r.get("dlt5_76")
        if espesor_code == "2":
            return r.get("dgt5_89"), r.get("dgt5_76")
    if r["tipo"] == 2:
        if espesor_code == "1":
            return r.get("blt5_89"), r.get("blt5_76")
        if espesor_code == "2":
            return r.get("bgt5_89"), r.get("bgt5_76")
    return None, None


def calc_ratings(row: dict):
    alt = next((a for a in ALTERACION_CAT if a["code"] == row.get("alteracion", "")), None)
    alt89 = alt["r89"] if alt else None
    alt76 = alt["r76"] if alt else None

    r1_89, r1_76 = get_relleno_rating(row.get("relleno1", ""), row.get("espesor", ""))
    r2_89, r2_76 = get_relleno_rating(row.get("relleno2", ""), row.get("espesor", ""))
    if r1_89 is not None and r2_89 is not None:
        rel89, rel76 = min(r1_89, r2_89), min(r1_76, r2_76)
    elif r1_89 is not None:
        rel89, rel76 = r1_89, r1_76
    elif r2_89 is not None:
        rel89, rel76 = r2_89, r2_76
    else:
        rel89, rel76 = None, None

    cont = next((c for c in CONTINUIDAD_CAT if c["code"] == row.get("continuidad", "")), None)
    cont89 = cont["r89"] if cont else None
    cont76 = cont["r76"] if cont else None

    aber = next((a for a in ABERTURA_CAT if a["code"] == row.get("abertura", "")), None)
    aber89 = aber["r89"] if aber else None
    aber76 = aber["r76"] if aber else None

    rug_pair = RUGOSIDAD_CAT.get(row.get("rugosidad", ""))
    rug89 = rug_pair[0] if rug_pair else None
    rug76 = rug_pair[1] if rug_pair else None

    vals89 = [alt89, rel89, cont89, aber89, rug89]
    vals76 = [alt76, rel76, cont76, aber76, rug76]
    val89 = sum(vals89) if all(v is not None for v in vals89) else None
    val76 = sum(vals76) if all(v is not None for v in vals76) else None

    return {
        "alt89": alt89, "alt76": alt76,
        "rel89": rel89, "rel76": rel76,
        "cont89": cont89, "cont76": cont76,
        "aber89": aber89, "aber76": aber76,
        "rug89": rug89, "rug76": rug76,
        "val89": val89, "val76": val76,
        "r1_89": r1_89, "r1_76": r1_76,
        "r2_89": r2_89, "r2_76": r2_76,
    }


def calc_summary(rows: list[dict]):
    familias = sorted(set(r["familia"] for r in rows))
    fam_stats = []
    jv = 0.0
    for f in familias:
        frows = [r for r in rows if r["familia"] == f]
        ratings = [calc_ratings(r) for r in frows]
        v89 = [rt["val89"] for rt in ratings if rt["val89"] is not None]
        v76 = [rt["val76"] for rt in ratings if rt["val76"] is not None]
        spacings = [float(r["espaciamiento"]) for r in frows
                    if r.get("espaciamiento", "") not in ("", None)
                    and float(r["espaciamiento"]) > 0]
        if spacings:
            avg_sp = sum(spacings) / len(spacings)
            jv += 1 / avg_sp
        fam_stats.append({
            "familia": f,
            "n": len(frows),
            "avg89": round(sum(v89) / len(v89), 4) if v89 else None,
            "avg76": round(sum(v76) / len(v76), 4) if v76 else None,
        })

    all_v89 = [calc_ratings(r)["val89"] for r in rows if calc_ratings(r)["val89"] is not None]
    all_v76 = [calc_ratings(r)["val76"] for r in rows if calc_ratings(r)["val76"] is not None]
    fam_avgs89 = [f["avg89"] for f in fam_stats if f["avg89"] is not None]
    fam_avgs76 = [f["avg76"] for f in fam_stats if f["avg76"] is not None]

    prom1_89 = round(sum(all_v89) / len(all_v89), 4) if all_v89 else None
    prom1_76 = round(sum(all_v76) / len(all_v76), 4) if all_v76 else None
    prom2_89 = round(sum(fam_avgs89) / len(fam_avgs89), 4) if fam_avgs89 else None
    prom2_76 = round(sum(fam_avgs76) / len(fam_avgs76), 4) if fam_avgs76 else None
    prom3_89 = round(min(fam_avgs89), 4) if fam_avgs89 else None
    prom3_76 = round(min(fam_avgs76), 4) if fam_avgs76 else None

    return {
        "prom1_89": prom1_89, "prom1_76": prom1_76,
        "prom2_89": prom2_89, "prom2_76": prom2_76,
        "prom3_89": prom3_89, "prom3_76": prom3_76,
        "jv": round(jv, 6),
        "fam_stats": fam_stats,
    }

# ─────────────────────────────────────────────────────────────────────────────
# ESTADO DE SESIÓN
# ─────────────────────────────────────────────────────────────────────────────

def init_state():
    if "header" not in st.session_state:
        st.session_state.header = {
            "td2": "", "ini_x": "", "ini_y": "", "ini_cota": "",
            "fin_x": "", "fin_y": "", "fin_cota": "",
            "largo": 10.0, "altura": 15.0,
            "dip_taled": "", "dipdir_tale": "", "dip_hole": "", "az_hole": "",
            "lito3": "", "lmt_m": "", "sector": "", "este": "",
            "fase": "", "nivel": "", "sect_geot": "", "mapeador": "",
            "intemperismo": "", "alteracion_zona": "",
        }
    if "rows" not in st.session_state:
        st.session_state.rows = [
            _empty_row("1"), _empty_row("1"), _empty_row("2"),
        ]


def _empty_row(familia="1"):
    import uuid
    return {
        "id": str(uuid.uuid4()),
        "familia": familia,
        "distancia": "",
        "dip": "",
        "dipdir": "",
        "tipo": "JN",
        "n_estr": 1,
        "abertura": "",
        "espesor": "",
        "continuidad": "",
        "espaciamiento": "",
        "n_extremos": "",
        "terminacion": "",
        "relleno1": "",
        "relleno2": "",
        "jrc": "",
        "rugosidad": "",
        "forma": "",
        "alteracion": "",
    }

# ─────────────────────────────────────────────────────────────────────────────
# EXPORTAR EXCEL
# ─────────────────────────────────────────────────────────────────────────────

def export_excel(header: dict, rows: list[dict]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Mapeo"

    orange_fill = PatternFill("solid", fgColor="F59E0B")
    pink_fill   = PatternFill("solid", fgColor="EC4899")
    yellow_fill = PatternFill("solid", fgColor="FEF3C7")
    gray_fill   = PatternFill("solid", fgColor="374151")
    bold = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    def cell(row, col, value, fill=None, font=None, align=center):
        c = ws.cell(row=row, column=col, value=value)
        if fill: c.fill = fill
        if font: c.font = font
        c.alignment = align
        c.border = thin
        return c

    # ── Cabecera del registro ────────────────────────────────────────────
    ws.merge_cells("A1:B1"); cell(1,1,"TD2", fill=gray_fill, font=bold)
    ws.merge_cells("C1:D1"); ws.cell(1,3).value = header["td2"]

    cell(1,5,"LARGO", fill=gray_fill, font=bold)
    ws.cell(1,6).value = header["largo"]
    cell(2,5,"ALTURA", fill=gray_fill, font=bold)
    ws.cell(2,6).value = header["altura"]

    cell(1,7,"Dip_taled", fill=gray_fill, font=bold)
    ws.cell(1,8).value = header["dip_taled"]
    cell(2,7,"DipDir_Tale", fill=gray_fill, font=bold)
    ws.cell(2,8).value = header["dipdir_tale"]

    cell(1,9,"LITO-3", fill=gray_fill, font=bold);  ws.cell(1,10).value = header["lito3"]
    cell(1,11,"LMT_M", fill=gray_fill, font=bold);  ws.cell(1,12).value = header["lmt_m"]
    cell(1,13,"SECTOR", fill=gray_fill, font=bold);  ws.cell(1,14).value = header["sector"]
    cell(2,9,"Mapeador", fill=gray_fill, font=bold); ws.cell(2,10).value = header["mapeador"]
    cell(2,11,"Intemp.", fill=gray_fill, font=bold); ws.cell(2,12).value = header["intemperismo"]
    cell(2,13,"Alt. Zona", fill=gray_fill, font=bold); ws.cell(2,14).value = header["alteracion_zona"]

    # ── Cabecera de tabla (fila 4) ───────────────────────────────────────
    COLS_BASE = [
        "ID", "Dist.(m)", "Dip(°)", "DipDir(°)", "Tipo", "N°Estr.",
        "Abertura", "Espesor", "Continuidad", "Espac.(m)", "N°Ext.",
        "Term.", "Relleno1", "Relleno2", "R1-89", "R2-89",
        "JRC", "Rugos.", "Forma", "Alter.",
    ]
    COLS_89 = ["Alt.89","Rel.89","Cont.89","Aber.89","Rug.89","Val.89"]
    COLS_76 = ["Alt.76","Rel.76","Cont.76","Aber.76","Rug.76","Val.76"]

    hdr_row = 4
    for ci, h in enumerate(COLS_BASE, start=1):
        cell(hdr_row, ci, h, fill=gray_fill, font=bold)
    start89 = len(COLS_BASE) + 1
    for ci, h in enumerate(COLS_89, start=start89):
        cell(hdr_row, ci, h, fill=PatternFill("solid", fgColor="B45309"), font=Font(bold=True, color="FFFFFF"))
    start76 = start89 + len(COLS_89)
    for ci, h in enumerate(COLS_76, start=start76):
        cell(hdr_row, ci, h, fill=PatternFill("solid", fgColor="9D174D"), font=Font(bold=True, color="FFFFFF"))

    # ── Filas de datos ───────────────────────────────────────────────────
    for ri, row in enumerate(rows, start=hdr_row + 1):
        rt = calc_ratings(row)
        base_vals = [
            row["familia"], row["distancia"], row["dip"], row["dipdir"],
            row["tipo"], row["n_estr"],
            next((a["label"] for a in ABERTURA_CAT if a["code"] == row.get("abertura","")), ""),
            next((e["label"] for e in ESPESOR_CAT if e["code"] == row.get("espesor","")), ""),
            next((c["label"] for c in CONTINUIDAD_CAT if c["code"] == row.get("continuidad","")), ""),
            row["espaciamiento"], row["n_extremos"], row["terminacion"],
            row["relleno1"], row["relleno2"],
            rt["r1_89"] if rt["r1_89"] is not None else "",
            rt["r2_89"] if rt["r2_89"] is not None else "",
            row["jrc"], row["rugosidad"], row["forma"],
            row["alteracion"],
        ]
        for ci, v in enumerate(base_vals, start=1):
            cell(ri, ci, v)
        rat89 = [rt["alt89"], rt["rel89"], rt["cont89"], rt["aber89"], rt["rug89"], rt["val89"]]
        rat76 = [rt["alt76"], rt["rel76"], rt["cont76"], rt["aber76"], rt["rug76"], rt["val76"]]
        for ci, v in enumerate(rat89, start=start89):
            cell(ri, ci, v if v is not None else "", fill=PatternFill("solid", fgColor="FEF3C7"))
        for ci, v in enumerate(rat76, start=start76):
            cell(ri, ci, v if v is not None else "", fill=PatternFill("solid", fgColor="FDF2F8"))

    # ── Resumen ──────────────────────────────────────────────────────────
    summ = calc_summary(rows)
    sr = len(rows) + hdr_row + 3
    cell(sr, 1, "RESUMEN", fill=gray_fill, font=bold)
    cell(sr+1, 1, "PROM 1 (todos)"); cell(sr+1, 2, summ["prom1_89"]); cell(sr+1, 3, summ["prom1_76"])
    cell(sr+2, 1, "PROM 2 (x familia)"); cell(sr+2, 2, summ["prom2_89"]); cell(sr+2, 3, summ["prom2_76"])
    cell(sr+3, 1, "PROM 3 (peor familia)"); cell(sr+3, 2, summ["prom3_89"]); cell(sr+3, 3, summ["prom3_76"])
    cell(sr+4, 1, "JV"); cell(sr+4, 2, summ["jv"])

    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 22)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

# ─────────────────────────────────────────────────────────────────────────────
# INTERFAZ PRINCIPAL
# ─────────────────────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Mapeo por Ventana Geomecánica",
    page_icon="⛏",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown("""
<style>
    .main .block-container { padding-top: 1rem; }
    div[data-testid="stDataFrame"] { font-size: 12px; }
    .stSelectbox label { font-size: 11px; }
    .stNumberInput label { font-size: 11px; }
    .stTextInput label { font-size: 11px; }
    .rating-89 { background-color: #92400e22; color: #f59e0b; font-weight: bold; }
    .rating-76 { background-color: #9d174d22; color: #ec4899; font-weight: bold; }
    h1 { font-size: 1.4rem !important; }
    h2 { font-size: 1.1rem !important; }
    h3 { font-size: 0.95rem !important; }
</style>
""", unsafe_allow_html=True)

init_state()
h = st.session_state.header

# ── Sidebar — Catálogos ──────────────────────────────────────────────────────
with st.sidebar:
    st.title("📚 Catálogos")

    with st.expander("Tipo de Estructura"):
        st.table(pd.DataFrame({
            "Código": TIPO_ESTRUCTURA_CODES,
            "Descripción": ["Junta","Estrato","Falla","Cizalla","Contacto"],
        }))

    with st.expander("Abertura"):
        df_ab = pd.DataFrame([
            {"Abertura": a["label"], "Rango mm": a["mm"], "Rating 89": a["r89"], "Rating 76": a["r76"]}
            for a in ABERTURA_CAT
        ])
        st.dataframe(df_ab, hide_index=True)

    with st.expander("Continuidad"):
        df_co = pd.DataFrame([
            {"Continuidad": c["label"], "Rating 89": c["r89"], "Rating 76": c["r76"]}
            for c in CONTINUIDAD_CAT
        ])
        st.dataframe(df_co, hide_index=True)

    with st.expander("Espesor"):
        st.table(pd.DataFrame({
            "Espesor (mm)": ["<5", ">5"],
            "Código": ["1", "2"],
        }))

    with st.expander("Relleno"):
        rows_rell = []
        for r in RELLENO_CAT:
            rows_rell.append({
                "Código": r["code"],
                "Sin(89)": r.get("sin89", ""),
                "D<5(89)": r.get("dlt5_89",""), "D>5(89)": r.get("dgt5_89",""),
                "B<5(89)": r.get("blt5_89",""), "B>5(89)": r.get("bgt5_89",""),
                "Sin(76)": r.get("sin76",""),
                "D<5(76)": r.get("dlt5_76",""), "D>5(76)": r.get("dgt5_76",""),
                "B<5(76)": r.get("blt5_76",""), "B>5(76)": r.get("bgt5_76",""),
            })
        st.dataframe(pd.DataFrame(rows_rell), hide_index=True)
        st.caption("D=Duro, B=Blando. Se toma el valor MÍNIMO entre Relleno 1 y Relleno 2.")

    with st.expander("Rugosidad"):
        df_rug = pd.DataFrame({
            "Perfil": list(RUGOSIDAD_CAT.keys()),
            "Rating 89": [v[0] for v in RUGOSIDAD_CAT.values()],
            "Rating 76": [v[1] for v in RUGOSIDAD_CAT.values()],
        })
        st.dataframe(df_rug, hide_index=True)

    with st.expander("Forma de Estructura"):
        st.table(pd.DataFrame({
            "Código": FORMA_ESTRUCTURA_CODES,
            "Forma": ["Plana","Curva","Ondulada","Escalonada","Irregular"],
        }))

    with st.expander("Meteorización / Alteración"):
        df_alt = pd.DataFrame([
            {"Código": a["code"], "Grado": a["label"].split("—")[1].strip(),
             "Rating 89": a["r89"], "Rating 76": a["r76"]}
            for a in ALTERACION_CAT
        ])
        st.dataframe(df_alt, hide_index=True)

    with st.expander("JRC — Rugosidad"):
        st.markdown("""
| Rango JRC | Perfil rugosidad |
|-----------|-----------------|
| 0–2       | 9               |
| 2–4       | 8               |
| 4–6       | 7               |
| 6–8       | 6               |
| 8–10      | 5               |
| 10–12     | 4               |
| 12–14     | 3               |
| 14–16     | 2               |
| 16–18     | 1               |
| 18–20     | 1               |
        """)

# ── Título ───────────────────────────────────────────────────────────────────
st.title("⛏ Mapeo por Ventana Geomecánica")
st.caption("Registro de discontinuidades con cálculo automático RMR'89 y RMR'76 — MMG / Las Bambas 2021")

# ── Sección 1: Datos de Registro ─────────────────────────────────────────────
st.markdown("---")
st.subheader("📋 Datos de Registro")

with st.container():
    c1, c2, c3, c4 = st.columns([1.5, 2, 2, 2])
    with c1:
        h["td2"] = st.text_input("TD2", value=h["td2"], key="td2")
    with c2:
        st.caption("Coordenadas INI (X / Y / Cota)")
        ci1, ci2, ci3 = st.columns(3)
        h["ini_x"]    = ci1.text_input("X", value=h["ini_x"], key="ini_x", label_visibility="collapsed")
        h["ini_y"]    = ci2.text_input("Y", value=h["ini_y"], key="ini_y", label_visibility="collapsed")
        h["ini_cota"] = ci3.text_input("Cota", value=h["ini_cota"], key="ini_cota", label_visibility="collapsed")
    with c3:
        st.caption("Coordenadas FIN (X / Y / Cota)")
        cf1, cf2, cf3 = st.columns(3)
        h["fin_x"]    = cf1.text_input("X", value=h["fin_x"], key="fin_x", label_visibility="collapsed")
        h["fin_y"]    = cf2.text_input("Y", value=h["fin_y"], key="fin_y", label_visibility="collapsed")
        h["fin_cota"] = cf3.text_input("Cota", value=h["fin_cota"], key="fin_cota", label_visibility="collapsed")
    with c4:
        h["mapeador"] = st.text_input("Mapeador", value=h["mapeador"], key="mapeador")

    col_set = st.columns(10)
    h["largo"]       = col_set[0].number_input("Largo (m)",     value=float(h["largo"]),     min_value=0.0, key="largo")
    h["altura"]      = col_set[1].number_input("Altura (m)",    value=float(h["altura"]),    min_value=0.0, key="altura")
    h["dip_taled"]   = col_set[2].text_input("Dip Taled (°)",   value=h["dip_taled"],        key="dip_taled")
    h["dipdir_tale"] = col_set[3].text_input("DipDir Tale (°)", value=h["dipdir_tale"],      key="dipdir_tale")
    h["dip_hole"]    = col_set[4].text_input("Dip Hole",        value=h["dip_hole"],         key="dip_hole")
    h["az_hole"]     = col_set[5].text_input("Az Hole",         value=h["az_hole"],          key="az_hole")
    h["lito3"]       = col_set[6].text_input("LITO-3",          value=h["lito3"],            key="lito3")
    h["lmt_m"]       = col_set[7].text_input("LMT_M",           value=h["lmt_m"],            key="lmt_m")
    h["sector"]      = col_set[8].text_input("Sector",          value=h["sector"],           key="sector")
    h["este"]        = col_set[9].text_input("Este",            value=h["este"],             key="este")

    col_set2 = st.columns(6)
    h["fase"]           = col_set2[0].text_input("Fase",            value=h["fase"],           key="fase")
    h["nivel"]          = col_set2[1].text_input("Nivel",           value=h["nivel"],          key="nivel")
    h["sect_geot"]      = col_set2[2].text_input("Sect. GEOT",      value=h["sect_geot"],      key="sect_geot")

    alt_labels_h = ["—"] + [a["label"] for a in ALTERACION_CAT if a["code"] != "-1"]
    alt_codes_h  = [""]  + [a["code"]  for a in ALTERACION_CAT if a["code"] != "-1"]
    cur_intemp = h.get("intemperismo", "")
    idx_intemp = alt_codes_h.index(cur_intemp) if cur_intemp in alt_codes_h else 0
    sel_intemp = col_set2[3].selectbox("Intemperismo", alt_labels_h, index=idx_intemp, key="intemperismo_sel")
    h["intemperismo"] = alt_codes_h[alt_labels_h.index(sel_intemp)]

    h["alteracion_zona"] = col_set2[4].text_input("Alteración de Zona", value=h["alteracion_zona"], key="alt_zona")

# ── Sección 2: Discontinuidades ───────────────────────────────────────────────
st.markdown("---")
st.subheader("⛏ Registro de Discontinuidades")

largo_max = float(h["largo"]) if h["largo"] else 999.0
rows = st.session_state.rows

# Botón para agregar filas
st.caption("Agregar fila a familia:")
add_cols = st.columns(10)
for fi in range(1, 10):
    if add_cols[fi - 1].button(str(fi), key=f"add_fam_{fi}", use_container_width=True):
        rows.append(_empty_row(str(fi)))
        st.rerun()

st.caption(f"**{len(rows)} filas** · Largo máximo: **{largo_max} m**")

# ─ Tabla editable ────────────────────────────────────────────────────────────
ABERTURA_LABELS = ["—"] + [a["label"] for a in ABERTURA_CAT]
ABERTURA_CODES  = [""]  + [a["code"]  for a in ABERTURA_CAT]
CONT_LABELS = ["—"] + [c["label"] for c in CONTINUIDAD_CAT]
CONT_CODES  = [""]  + [c["code"]  for c in CONTINUIDAD_CAT]
ESPESOR_LABELS = ["—", "<5 mm", ">5 mm"]
ESPESOR_CODES  = ["", "1", "2"]
RUGOSIDAD_LABELS = ["—"] + [str(k) for k in RUGOSIDAD_CAT.keys()]
RUGOSIDAD_CODES  = [""]  + [str(k) for k in RUGOSIDAD_CAT.keys()]
RELLENO_LABELS = ["—"] + [r["code"] for r in RELLENO_CAT]
RELLENO_CODES  = [""]  + [r["code"] for r in RELLENO_CAT]

to_delete = []
for i, row in enumerate(rows):
    with st.expander(
        f"Fila {i+1}  |  Familia {row['familia']}  |  {row.get('tipo','JN')}  "
        f"Dip {row.get('dip','—')} / {row.get('dipdir','—')}",
        expanded=(i == len(rows) - 1)
    ):
        rc = st.columns([1, 1, 1, 1, 1.5, 1])
        row["familia"] = rc[0].selectbox("Familia", [str(x) for x in range(1,10)],
            index=int(row["familia"])-1, key=f"fam_{i}")

        dist_val = rc[1].text_input("Distancia (m)", value=row["distancia"], key=f"dist_{i}")
        row["distancia"] = dist_val
        try:
            if float(dist_val) > largo_max:
                rc[1].error(f"⚠ > {largo_max} m")
        except (ValueError, TypeError):
            pass

        row["dip"]    = rc[2].text_input("Dip (°)",    value=row["dip"],    key=f"dip_{i}")
        row["dipdir"] = rc[3].text_input("DipDir (°)", value=row["dipdir"], key=f"dipdir_{i}")

        tipo_idx = TIPO_ESTRUCTURA_CODES.index(row["tipo"]) if row["tipo"] in TIPO_ESTRUCTURA_CODES else 0
        row["tipo"] = rc[4].selectbox("Tipo Estructura",
            TIPO_ESTRUCTURA_CODES, index=tipo_idx, key=f"tipo_{i}")
        row["n_estr"] = rc[5].number_input("N° Estr.", value=int(row.get("n_estr", 1)),
            min_value=1, key=f"nestr_{i}")

        rc2 = st.columns([2, 1.5, 2, 1, 1, 1.5])
        aber_idx = ABERTURA_CODES.index(row["abertura"]) if row["abertura"] in ABERTURA_CODES else 0
        row["abertura"] = ABERTURA_CODES[
            rc2[0].selectbox("Abertura", ABERTURA_LABELS, index=aber_idx, key=f"aber_{i}")
            and ABERTURA_LABELS.index(rc2[0].selectbox("Abertura", ABERTURA_LABELS, index=aber_idx, key=f"aber2_{i}"))
        ] if False else ABERTURA_CODES[aber_idx]  # simplified below

        # Simpler approach for dropdowns
        sel_aber = rc2[0].selectbox("Abertura", ABERTURA_LABELS,
            index=ABERTURA_CODES.index(row["abertura"]) if row["abertura"] in ABERTURA_CODES else 0,
            key=f"aber_sel_{i}")
        row["abertura"] = ABERTURA_CODES[ABERTURA_LABELS.index(sel_aber)]

        sel_esp = rc2[1].selectbox("Espesor",   ESPESOR_LABELS,
            index=ESPESOR_CODES.index(row["espesor"]) if row["espesor"] in ESPESOR_CODES else 0,
            key=f"esp_sel_{i}")
        row["espesor"] = ESPESOR_CODES[ESPESOR_LABELS.index(sel_esp)]

        sel_cont = rc2[2].selectbox("Continuidad", CONT_LABELS,
            index=CONT_CODES.index(row["continuidad"]) if row["continuidad"] in CONT_CODES else 0,
            key=f"cont_sel_{i}")
        row["continuidad"] = CONT_CODES[CONT_LABELS.index(sel_cont)]

        row["espaciamiento"] = rc2[3].text_input("Espac. (m)", value=row["espaciamiento"], key=f"espac_{i}")
        row["n_extremos"]    = rc2[4].text_input("N° Ext.",   value=row["n_extremos"],    key=f"next_{i}")

        sel_term = rc2[5].selectbox("Terminación", ["—"]+TERMINACION_CODES,
            index=(["—"]+TERMINACION_CODES).index(row["terminacion"])
            if row["terminacion"] in ["—"]+TERMINACION_CODES else 0,
            key=f"term_sel_{i}")
        row["terminacion"] = "" if sel_term == "—" else sel_term

        rc3 = st.columns([1.5, 1.5, 1, 1, 1.5, 1.5])
        sel_r1 = rc3[0].selectbox("Relleno 1", RELLENO_LABELS,
            index=RELLENO_CODES.index(row["relleno1"]) if row["relleno1"] in RELLENO_CODES else 0,
            key=f"r1_sel_{i}")
        row["relleno1"] = RELLENO_CODES[RELLENO_LABELS.index(sel_r1)]

        sel_r2 = rc3[1].selectbox("Relleno 2", RELLENO_LABELS,
            index=RELLENO_CODES.index(row["relleno2"]) if row["relleno2"] in RELLENO_CODES else 0,
            key=f"r2_sel_{i}")
        row["relleno2"] = RELLENO_CODES[RELLENO_LABELS.index(sel_r2)]

        row["jrc"] = rc3[2].text_input("JRC (0–20)", value=row["jrc"], key=f"jrc_{i}")

        sel_rug = rc3[3].selectbox("Rugosidad", RUGOSIDAD_LABELS,
            index=RUGOSIDAD_CODES.index(row["rugosidad"]) if row["rugosidad"] in RUGOSIDAD_CODES else 0,
            key=f"rug_sel_{i}")
        row["rugosidad"] = RUGOSIDAD_CODES[RUGOSIDAD_LABELS.index(sel_rug)]

        sel_forma = rc3[4].selectbox("Forma", ["—"]+FORMA_ESTRUCTURA_CODES,
            index=(["—"]+FORMA_ESTRUCTURA_CODES).index(row["forma"])
            if row["forma"] in ["—"]+FORMA_ESTRUCTURA_CODES else 0,
            key=f"forma_sel_{i}")
        row["forma"] = "" if sel_forma == "—" else sel_forma

        sel_alt = rc3[5].selectbox("Alteración", ALT_LABELS,
            index=ALT_CODES.index(row["alteracion"]) if row["alteracion"] in ALT_CODES else 0,
            key=f"alt_sel_{i}")
        row["alteracion"] = ALT_CODES[ALT_LABELS.index(sel_alt)]

        # Ratings calculados
        rt = calc_ratings(row)
        r1_89_v, _ = get_relleno_rating(row["relleno1"], row["espesor"])
        r2_89_v, _ = get_relleno_rating(row["relleno2"], row["espesor"])

        rv_cols = st.columns(14)
        rv_cols[0].metric("Val Rell.1 (89)", r1_89_v if r1_89_v is not None else "—")
        rv_cols[1].metric("Val Rell.2 (89)", r2_89_v if r2_89_v is not None else "—")

        st.markdown("**Condición de Discontinuidades — RMR'89** 🟠 &nbsp; | &nbsp; **RMR'76** 🩷")
        r89_cols = st.columns(6)
        r89_cols[0].metric("Alt.89",  rt["alt89"]  if rt["alt89"]  is not None else "—")
        r89_cols[1].metric("Rell.89", rt["rel89"]  if rt["rel89"]  is not None else "—")
        r89_cols[2].metric("Cont.89", rt["cont89"] if rt["cont89"] is not None else "—")
        r89_cols[3].metric("Aber.89", rt["aber89"] if rt["aber89"] is not None else "—")
        r89_cols[4].metric("Rug.89",  rt["rug89"]  if rt["rug89"]  is not None else "—")
        r89_cols[5].metric("**VALOR 89**", rt["val89"] if rt["val89"] is not None else "—")

        r76_cols = st.columns(6)
        r76_cols[0].metric("Alt.76",  rt["alt76"]  if rt["alt76"]  is not None else "—")
        r76_cols[1].metric("Rell.76", rt["rel76"]  if rt["rel76"]  is not None else "—")
        r76_cols[2].metric("Cont.76", rt["cont76"] if rt["cont76"] is not None else "—")
        r76_cols[3].metric("Aber.76", rt["aber76"] if rt["aber76"] is not None else "—")
        r76_cols[4].metric("Rug.76",  rt["rug76"]  if rt["rug76"]  is not None else "—")
        r76_cols[5].metric("**VALOR 76**", rt["val76"] if rt["val76"] is not None else "—")

        if st.button("🗑 Eliminar esta fila", key=f"del_{i}", type="secondary"):
            to_delete.append(i)

if to_delete:
    for idx in sorted(to_delete, reverse=True):
        rows.pop(idx)
    st.rerun()

# ── Sección 3: Resumen ───────────────────────────────────────────────────────
st.markdown("---")
st.subheader("📊 Resumen")

if rows:
    summ = calc_summary(rows)
    sc = st.columns(4)
    sc[0].markdown("**Promedios**")
    sc[0].metric("PROM 1 — todos",       f"{summ['prom1_89']} / {summ['prom1_76']}" if summ["prom1_89"] else "—")
    sc[0].metric("PROM 2 — x familia",   f"{summ['prom2_89']} / {summ['prom2_76']}" if summ["prom2_89"] else "—")
    sc[0].metric("PROM 3 — peor familia",f"{summ['prom3_89']} / {summ['prom3_76']}" if summ["prom3_89"] else "—")
    sc[0].caption("Format: RMR'89 / RMR'76")

    sc[1].markdown("**Índice Volumétrico**")
    sc[1].metric("JV (jts/m³)", summ["jv"] if summ["jv"] else "—")

    sc[2].markdown("**Por Familia**")
    if summ["fam_stats"]:
        df_fam = pd.DataFrame(summ["fam_stats"]).rename(columns={
            "familia":"Fam","n":"N","avg89":"Avg 89","avg76":"Avg 76"})
        sc[2].dataframe(df_fam, hide_index=True)
else:
    st.info("Agrega filas para ver el resumen.")

# ── Sección 4: Exportar ───────────────────────────────────────────────────────
st.markdown("---")
st.subheader("💾 Exportar")

ec = st.columns(3)

# JSON
json_data = json.dumps({"header": h, "rows": rows, "fecha": str(date.today())}, indent=2, ensure_ascii=False)
ec[0].download_button(
    "⬇ Descargar JSON",
    data=json_data.encode("utf-8"),
    file_name=f"mapeo_{h.get('td2','sin_id')}_{date.today()}.json",
    mime="application/json",
    use_container_width=True,
)

# CSV
csv_cols = ["familia","distancia","dip","dipdir","tipo","n_estr","abertura","espesor",
            "continuidad","espaciamiento","n_extremos","terminacion","relleno1","relleno2",
            "jrc","rugosidad","forma","alteracion"]
df_exp = pd.DataFrame([{c: r.get(c,"") for c in csv_cols} for r in rows])
ec[1].download_button(
    "⬇ Descargar CSV",
    data=df_exp.to_csv(index=False).encode("utf-8"),
    file_name=f"mapeo_{h.get('td2','sin_id')}_{date.today()}.csv",
    mime="text/csv",
    use_container_width=True,
)

# Excel
try:
    excel_bytes = export_excel(h, rows)
    ec[2].download_button(
        "⬇ Descargar Excel (.xlsx)",
        data=excel_bytes,
        file_name=f"mapeo_{h.get('td2','sin_id')}_{date.today()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )
except Exception as e:
    ec[2].warning(f"Excel no disponible: {e}\nInstala openpyxl: pip install openpyxl")

# ── Limpiar ───────────────────────────────────────────────────────────────────
st.markdown("---")
if st.button("🗑 Limpiar todos los datos", type="secondary"):
    for k in list(st.session_state.keys()):
        del st.session_state[k]
    st.rerun()
